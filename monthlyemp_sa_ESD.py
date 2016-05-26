#Import the modules/packages

import os
import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook
import urllib
import datetime
import time

#***"ATTN: Please seletct the first three letter of the month of interest for monthly data assembly:"***
active_month = "Apr"

#url for seasonally adjusted data:

esd_sa_url = "https://fortress.wa.gov/esd/employmentdata/docs/economic-reports/wa-historical-employment-seasonally-adjusted.xlsx"

#define the excel file name:

sa_filename = esd_sa_url[-49:-5]+"_"+time.strftime("%m%d%Y")+esd_sa_url[-5:]


#download the data
def download_esd_sa():
    start = time.time()
    esdfile = urllib.URLopener()
    esdfile.retrieve(esd_sa_url, os.getcwd()+"/"+ sa_filename)
    end = time.time()
    processtime = end - start
    print "\nData download completed. Download Time: " + "{0: .3}".format(processtime) + " Sec."


# Check the number of worksheet in the .xlsx file & list the name of the worksheets
#included in the analysis

def check_worksheet():
    esd_sa_file = pd.ExcelFile(os.getcwd()+"/"+ sa_filename)
    #get the list of worksheets from the excelfile
    esd_sa_ws_master = {sheet_name: esd_sa_file.parse(sheet_name) 
              for sheet_name in esd_sa_file.sheet_names}

    a_list = []
    count = 0
    count_total = 0
    for i in esd_sa_ws_master:
        if i == "Seattle MSA":
            a_list.append(i)
            count = count+1
            count_total = count_total+1
        elif i == "Bremerton MSA":
            a_list.append(i)
            count = count+1
            count_total = count_total+1
        elif i == "Tacoma MSA":
            a_list.append(i)
            count = count+1
            count_total = count_total+1
        else:
            count_total = count_total+1

    print "\nList of worksheets to be included in the analysis: "

    count_ws = 0
    for x in sorted(a_list):
        count_ws = count_ws+1 
        print str(count_ws)+". "+ x

    print "Total "+str(count)+" "+"out of "+ str(count_total)+ " worksheets contain data for the PSRC region."


def sa_kin_sno():
    #read the required worksheets into dataframes
    df_sa_kin_sno = pd.read_excel(os.getcwd()+"/"+ sa_filename, sheetname="Seattle MSA", header = 1, index_col=[1])

    #transpose the dataframe
    df_sa_kin_sno = df_sa_kin_sno.transpose()

    #transposing caused the date column to be the index column, so
    #extracting the date information to another column named "date_stamp"

    df_sa_kin_sno['date_stamp'] = df_sa_kin_sno.index

    #deleting the first data row (not the header) containing NAICS info 
    #entirely for formatting purposes 

    df_sa_kin_sno = df_sa_kin_sno.ix[1:]

    #resetting the index of the dataframe
    #as we already copied the date info into another column, nothing to worry

    df_sa_kin_sno = df_sa_kin_sno.reset_index(drop =True)

    #renaming the index colum from "NAICS Industy" to "id" 

    df_sa_kin_sno.columns.names = ["id"]

    #get a column for year and one for month (note the foramting for month)
    df_sa_kin_sno["year"] = df_sa_kin_sno["date_stamp"].dt.year
    df_sa_kin_sno["month"] = df_sa_kin_sno["date_stamp"].dt.strftime("%b")

    #get the month column to string (not essential)
    df_sa_kin_sno["month"] = df_sa_kin_sno["month"].astype("str")

    #define sectors for King and Snohomish Counties:

    #get the list of column names, any of the options below works just fine
    # list(df_sa_kin_sno) or list(df_sa_kin_sno.columns.values) works too
    kin_sno_list = df_sa_kin_sno.columns.values.tolist()

    #define PSRC sectors by adding columns for each of the eight sectors and total nonfarm jobs:
    df_sa_kin_sno["King & Snohomish: Total NonFarm"] = df_sa_kin_sno["Total Nonfarm"] 
    df_sa_kin_sno["King & Snohomish: Consruction"] = df_sa_kin_sno[kin_sno_list[3]] + df_sa_kin_sno[kin_sno_list[4]]
    df_sa_kin_sno["King & Snohomish: FIRE"] = df_sa_kin_sno[kin_sno_list[33]]
    df_sa_kin_sno["King & Snohomish: Manufacturing"] = df_sa_kin_sno[kin_sno_list[8]]
    df_sa_kin_sno["King & Snohomish: Retail"] = df_sa_kin_sno[kin_sno_list[20]]
    df_sa_kin_sno["King & Snohomish: Service"] = df_sa_kin_sno[kin_sno_list[17]] - df_sa_kin_sno[kin_sno_list[18]] - df_sa_kin_sno[kin_sno_list[33]] - df_sa_kin_sno[kin_sno_list[62]]                                                     
    df_sa_kin_sno["King & Snohomish: WTU"] = df_sa_kin_sno[kin_sno_list[18]] - df_sa_kin_sno[kin_sno_list[20]]
    df_sa_kin_sno["King & Snohomish: Government"] = df_sa_kin_sno[kin_sno_list[62]] - df_sa_kin_sno[kin_sno_list[65]] - df_sa_kin_sno[kin_sno_list[67]]
    df_sa_kin_sno["King & Snohomish: Education"] = df_sa_kin_sno[kin_sno_list[65]] + df_sa_kin_sno[kin_sno_list[67]]

    #delete columns that we don't need anymore (all except the year, month and columns created for sectors)
    #df_sa_kin_sno = df_sa_kin_sno.drop(df_sa_kin_sno.columns[range(69)], axis = 1) works for this operation too!
    df_sa_kin_sno = df_sa_kin_sno.drop(df_sa_kin_sno.columns[:69], axis = 1)

    #ATTENTION: """just for this iteration, won't need for future runs""", deleting the row that contains Apr16 info for Seattle MSA
    #df_sa_kin_sno = df_sa_kin_sno.drop(df_sa_kin_sno.index[[315]])

    #putting in a check-point in place to make sure the sector columns add up to the total jobs:
    #df_sa_kit["Kitsap: Total NonFarm"].isin(df_sa_kit[kit_col_list[3:]].sum(axis=1)).all() == True.... works too for this step.

    if df_sa_kin_sno["King & Snohomish: Total NonFarm"].isin(df_sa_kin_sno.ix[:,3:].sum(axis=1)).all() == True:
        print "\nChecked: King & Snohomish Counties- All industry sectors add up to total NonFarm jobs..."
    else:
        print "\nERROR!!"
    return df_sa_kin_sno

    
def sa_kit():
    df_sa_kit = pd.read_excel(os.getcwd()+"/"+ sa_filename, sheetname="Bremerton MSA", header = 1, index_col=[1])
    df_sa_kit = df_sa_kit.transpose()
    df_sa_kit['date_stamp'] = df_sa_kit.index
    df_sa_kit = df_sa_kit.ix[1:]
    df_sa_kit = df_sa_kit.reset_index(drop =True)
    df_sa_kit.columns.names = ["id"]
    df_sa_kit["year"] = df_sa_kit["date_stamp"].dt.year
    df_sa_kit["month"] = df_sa_kit["date_stamp"].dt.strftime("%b")

    #get the month column to string (not essential)
    df_sa_kit["month"] = df_sa_kit["month"].astype("str")

    kit_list = df_sa_kit.columns.values.tolist()

    df_sa_kit["Kitsap: Total NonFarm"] = df_sa_kit["Total Nonfarm"] 
    df_sa_kit["Kitsap: Consruction"] = df_sa_kit[kit_list[3]]
    #add new column for FIRE with zero as values with index aligned with the dataframe
    df_sa_kit["Kitsap: FIRE"] = pd.Series([0 for x in range(len(df_sa_kit.index))], index=df_sa_kit.index)
    df_sa_kit["Kitsap: Manufacturing"] = df_sa_kit[kit_list[4]]
    df_sa_kit["Kitsap: Retail"] = df_sa_kit[kit_list[8]]
    df_sa_kit["Kitsap: Service"] = df_sa_kit[kit_list[5]] - df_sa_kit[kit_list[7]] - df_sa_kit[kit_list[11]]
    df_sa_kit["Kitsap: WTU"] = df_sa_kit[kit_list[7]] - df_sa_kit[kit_list[8]]
    df_sa_kit["Kitsap: Government"] = df_sa_kit[kit_list[11]]
    df_sa_kit["Kitsap: Education"] = pd.Series([0 for x in range(len(df_sa_kit.index))], index=df_sa_kit.index)

    df_sa_kit = df_sa_kit.drop(df_sa_kit.columns[:16], axis = 1)

    if df_sa_kit["Kitsap: Total NonFarm"].isin(df_sa_kit.ix[:,3:].sum(axis=1)).all() == True:
        print "Checked: Kitsap County- All industry sectors add up to total NonFarm jobs..."
    else:
        print "\nERROR!!"
    return df_sa_kit



def sa_pie():
    df_sa_pie = pd.read_excel(os.getcwd()+"/"+ sa_filename, sheetname="Tacoma MSA", header = 1, index_col=[1])
    df_sa_pie = df_sa_pie.transpose()
    df_sa_pie['date_stamp'] = df_sa_pie.index
    df_sa_pie = df_sa_pie.ix[1:]
    df_sa_pie = df_sa_pie.reset_index(drop =True)
    df_sa_pie.columns.names = ["id"]
    df_sa_pie["year"] = df_sa_pie["date_stamp"].dt.year
    df_sa_pie["month"] = df_sa_pie["date_stamp"].dt.strftime("%b")

    #get the month column to string (not essential)
    df_sa_pie["month"] = df_sa_pie["month"].astype("str")

    pie_list = df_sa_pie.columns.values.tolist()

    df_sa_pie["Pierce: Total NonFarm"] = df_sa_pie["Total Nonfarm"] 
    df_sa_pie["Pierce: Consruction"] = df_sa_pie[pie_list[3]] + df_sa_pie[pie_list[4]]
    df_sa_pie["Pierce: FIRE"] = df_sa_pie[pie_list[15]]
    df_sa_pie["Pierce: Manufacturing"] = df_sa_pie[pie_list[6]]
    df_sa_pie["Pierce: Retail"] = df_sa_pie[pie_list[10]]
    df_sa_pie["Pierce: Service"] = df_sa_pie[pie_list[7]] - df_sa_pie[pie_list[8]] - df_sa_pie[pie_list[15]] - df_sa_pie[pie_list[25]]                                                     
    df_sa_pie["Pierce: WTU"] = df_sa_pie[pie_list[8]] - df_sa_pie[pie_list[10]]
    df_sa_pie["Pierce: Government"] = df_sa_pie[pie_list[25]] - df_sa_pie[pie_list[28]] - df_sa_pie[pie_list[30]]
    df_sa_pie["Pierce: Education"] = df_sa_pie[pie_list[28]] + df_sa_pie[pie_list[30]]

    df_sa_pie = df_sa_pie.drop(df_sa_pie.columns[:32], axis = 1)    

    if df_sa_pie["Pierce: Total NonFarm"].isin(df_sa_pie.ix[:,3:].sum(axis=1)).all() == True:
        print "Checked: Pierce County- All industry sectors add up to total NonFarm jobs..."
    else:
        print "\nERROR!!"
    return df_sa_pie



#merge the three dataframes to produce a regional dataframe
def sa_region():
    df_sa_psrc = pd.merge(pd.merge(df_sa_kin_sno, df_sa_pie, on = ["year", "month"]), df_sa_kit, on = ["year", "month"])
    
    
    #define regional sectors and add columns to a new dataframe

    df_sa_region  = df_sa_psrc
    region_list = df_sa_region.columns.values.tolist()

    df_sa_region["Region: Construction"] = df_sa_region[region_list[3]] + df_sa_region[region_list[12]] + df_sa_region[region_list[21]]
    df_sa_region["Region: FIRE"] = df_sa_region[region_list[4]] + df_sa_region[region_list[13]] + df_sa_region[region_list[22]]
    df_sa_region["Region: Manufacturing"] = df_sa_region[region_list[5]] + df_sa_region[region_list[14]] + df_sa_region[region_list[23]]
    df_sa_region["Region: Retail"] = df_sa_region[region_list[6]] + df_sa_region[region_list[15]] + df_sa_region[region_list[24]]
    df_sa_region["Region: Service"] = df_sa_region[region_list[7]] + df_sa_region[region_list[16]] + df_sa_region[region_list[25]]
    df_sa_region["Region: WTU"] = df_sa_region[region_list[8]] + df_sa_region[region_list[17]] + df_sa_region[region_list[26]]
    df_sa_region["Region: Education"] = df_sa_region[region_list[10]] + df_sa_region[region_list[19]] + df_sa_region[region_list[28]]
    df_sa_region["Region: Government"] = df_sa_region[region_list[9]] + df_sa_region[region_list[18]] + df_sa_region[region_list[27]]
    df_sa_region["Region: Total NonFarm"] = df_sa_region[region_list[2]] + df_sa_region[region_list[11]] + df_sa_region[region_list[20]]
    
    
    print "Checked: Regional dataframe completed..."
    
    return df_sa_region


#delete unnecessary worksheets from the excel file

def delete_ws():
    wb_sa = load_workbook(os.getcwd()+"/"+ sa_filename)

    del_sheets = []

    for i in wb_sa.get_sheet_names():
        if i == "Seattle MSA":
            pass
        elif i == "Tacoma MSA":
            pass
        elif i == "Bremerton MSA":
            pass
        elif i == "Washington State":
            pass
        else:
            del_sheets.append(i)
    #print del_sheets
    #now delete the bad sheets
    [wb_sa.remove_sheet(wb_sa.get_sheet_by_name(sheet)) for sheet in del_sheets]
    #if you get a ValueError, this means you are incorrectly requesting a sheet

    wb_sa.save(os.getcwd()+"/"+ sa_filename)
    
    print "\nWorksheets not included in the processing been deleted..."


def data_to_workbook():
    sa_book = load_workbook(os.getcwd()+"/"+ sa_filename)
    sa_writer = pd.ExcelWriter(os.getcwd()+"/"+ sa_filename, engine='openpyxl') 
    sa_writer.book = sa_book
    sa_writer.sheets = dict((ws.title, ws) for ws in sa_book.worksheets)

    df_sa_region_total.to_excel(sa_writer, "Region_Master", index =False)

    df_sa_region.to_excel(sa_writer, "Region_Sector", index =False)

    df_sa_region_total_CM.to_excel(sa_writer, "Region_CMonth", index =False)

    df_sa_region_sector_CM.to_excel(sa_writer, "Region_Sec_CMonth", index =False)

    sa_writer.save()
    
    print "\nProcessed data been exported to the excel file located at the working directory: " + "\n" + os.getcwd() +"\n" + "\nNew Worksheets: \n1. Region_Master \n2. Region_Sector \n3. Region_CMonth \n4. Region_Sec_CMonth "



process_start = time.time()
    
download_esd_sa()
check_worksheet()

df_sa_kin_sno = sa_kin_sno()
df_sa_kit = sa_kit()
df_sa_pie = sa_pie()
df_sa_region = sa_region()

  

#Query: just the three MSA and regional total
df_sa_region_total = df_sa_region[[0,1, 2, 11,20,-1]]
#Query: just the three MSA and regional total for current month
df_sa_region_total_CM = df_sa_region[[0,1, 2, 11,20,-1]][df_sa_region.month== active_month]
#Query: just the three MSA and regional sectors for current month
df_sa_region_sector_CM = df_sa_region[df_sa_region.month== active_month]    

delete_ws() 
    
data_to_workbook()

process_end = time.time()

processtime = process_end - process_start    

print "\nProcessing Completed! \nTotal Processing Time: " + "{0: .3}".format(processtime) + " Sec."