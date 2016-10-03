#Import the modules/packages

import os
import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook
import urllib
import datetime
import time

import plotly
import plotly.plotly as py
from plotly.tools import FigureFactory as FF
from plotly import tools
import plotly.graph_objs as go

#***"ATTN: Please seletct the first three letter of the month of interest for monthly data assembly:"***
active_month = "Jun"
hist_year = 2010

#url for seasonally adjusted data:

esd_sa_url = "https://fortress.wa.gov/esd/employmentdata/docs/economic-reports/wa-historical-employment-seasonally-adjusted.xlsx"

#define the excel file name:

sa_filename = esd_sa_url[-49:-5]+"_"+time.strftime("%m%d%Y")+esd_sa_url[-5:]


#download the data
def download_esd_sa():
    start = time.time()
    esdfile = urllib.URLopener()
    esdfile.retrieve(esd_sa_url, os.getcwd()+"/"+ sa_filename)
    end = time.time()
    processtime = end - start
    print "\nData download completed. Download Time: " + "{0: .3}".format(processtime) + " Sec."


# Check the number of worksheet in the .xlsx file & list the name of the worksheets
#included in the analysis

def check_worksheet():
    esd_sa_file = pd.ExcelFile(os.getcwd()+"/"+ sa_filename)
    #get the list of worksheets from the excelfile
    esd_sa_ws_master = {sheet_name: esd_sa_file.parse(sheet_name) 
              for sheet_name in esd_sa_file.sheet_names}

    a_list = []
    count = 0
    count_total = 0
    for i in esd_sa_ws_master:
        if i == "Seattle MSA":
            a_list.append(i)
            count = count+1
            count_total = count_total+1
        elif i == "Bremerton MSA":
            a_list.append(i)
            count = count+1
            count_total = count_total+1
        elif i == "Tacoma MSA":
            a_list.append(i)
            count = count+1
            count_total = count_total+1
        else:
            count_total = count_total+1

    print "\nList of worksheets to be included in the analysis: "

    count_ws = 0
    for x in sorted(a_list):
        count_ws = count_ws+1 
        print str(count_ws)+". "+ x

    print "Total "+str(count)+" "+"out of "+ str(count_total)+ " worksheets contain data for the PSRC region."


def sa_kin_sno():
    #read the required worksheets into dataframes
    df_sa_kin_sno = pd.read_excel(os.getcwd()+"/"+ sa_filename, sheetname="Seattle MSA", header = 1, index_col=[1])

    #transpose the dataframe
    df_sa_kin_sno = df_sa_kin_sno.transpose()

    #transposing caused the date column to be the index column, so
    #extracting the date information to another column named "date_stamp"

    df_sa_kin_sno['date_stamp'] = df_sa_kin_sno.index

    #deleting the first data row (not the header) containing NAICS info 
    #entirely for formatting purposes 

    df_sa_kin_sno = df_sa_kin_sno.ix[1:]

    #resetting the index of the dataframe
    #as we already copied the date info into another column, nothing to worry

    df_sa_kin_sno = df_sa_kin_sno.reset_index(drop =True)

    #renaming the index colum from "NAICS Industy" to "id" 

    df_sa_kin_sno.columns.names = ["id"]

    #get a column for year and one for month (note the foramting for month)
    df_sa_kin_sno["year"] = df_sa_kin_sno["date_stamp"].dt.year
    df_sa_kin_sno["month"] = df_sa_kin_sno["date_stamp"].dt.strftime("%b")

    #get the month column to string (not essential)
    df_sa_kin_sno["month"] = df_sa_kin_sno["month"].astype("str")

    #define sectors for King and Snohomish Counties:

    #get the list of column names, any of the options below works just fine
    # list(df_sa_kin_sno) or list(df_sa_kin_sno.columns.values) works too
    kin_sno_list = df_sa_kin_sno.columns.values.tolist()

    #define PSRC sectors by adding columns for each of the eight sectors and total nonfarm jobs:
    df_sa_kin_sno["King & Snohomish: Total NonFarm"] = df_sa_kin_sno["Total Nonfarm"] 
    df_sa_kin_sno["King & Snohomish: Const./Res"] = df_sa_kin_sno[kin_sno_list[3]] + df_sa_kin_sno[kin_sno_list[4]]
    df_sa_kin_sno["King & Snohomish: FIRE"] = df_sa_kin_sno[kin_sno_list[33]]
    df_sa_kin_sno["King & Snohomish: Manufacturing"] = df_sa_kin_sno[kin_sno_list[8]]
    df_sa_kin_sno["King & Snohomish: Retail"] = df_sa_kin_sno[kin_sno_list[20]]
    df_sa_kin_sno["King & Snohomish: Service"] = df_sa_kin_sno[kin_sno_list[17]] - df_sa_kin_sno[kin_sno_list[18]] - df_sa_kin_sno[kin_sno_list[33]] - df_sa_kin_sno[kin_sno_list[62]]                                                     
    df_sa_kin_sno["King & Snohomish: WTU"] = df_sa_kin_sno[kin_sno_list[18]] - df_sa_kin_sno[kin_sno_list[20]]
    df_sa_kin_sno["King & Snohomish: Government"] = df_sa_kin_sno[kin_sno_list[62]] - df_sa_kin_sno[kin_sno_list[65]] - df_sa_kin_sno[kin_sno_list[67]]
    df_sa_kin_sno["King & Snohomish: Education"] = df_sa_kin_sno[kin_sno_list[65]] + df_sa_kin_sno[kin_sno_list[67]]

    #delete columns that we don't need anymore (all except the year, month and columns created for sectors)
    #df_sa_kin_sno = df_sa_kin_sno.drop(df_sa_kin_sno.columns[range(69)], axis = 1) works for this operation too!
    df_sa_kin_sno = df_sa_kin_sno.drop(df_sa_kin_sno.columns[:69], axis = 1)

    #ATTENTION: """just for this iteration, won't need for future runs""", deleting the row that contains Apr16 info for Seattle MSA
    #df_sa_kin_sno = df_sa_kin_sno.drop(df_sa_kin_sno.index[[315]])

    #putting in a check-point in place to make sure the sector columns add up to the total jobs:
    #df_sa_kit["Kitsap: Total NonFarm"].isin(df_sa_kit[kit_col_list[3:]].sum(axis=1)).all() == True.... works too for this step.

    if df_sa_kin_sno["King & Snohomish: Total NonFarm"].isin(df_sa_kin_sno.ix[:,3:].sum(axis=1)).all() == True:
        print "\nChecked: King & Snohomish Counties- All industry sectors add up to total NonFarm jobs..."
    else:
        print "\nERROR!!"
    return df_sa_kin_sno

    
def sa_kit():
    df_sa_kit = pd.read_excel(os.getcwd()+"/"+ sa_filename, sheetname="Bremerton MSA", header = 1, index_col=[1])
    df_sa_kit = df_sa_kit.transpose()
    df_sa_kit['date_stamp'] = df_sa_kit.index
    df_sa_kit = df_sa_kit.ix[1:]
    df_sa_kit = df_sa_kit.reset_index(drop =True)
    df_sa_kit.columns.names = ["id"]
    df_sa_kit["year"] = df_sa_kit["date_stamp"].dt.year
    df_sa_kit["month"] = df_sa_kit["date_stamp"].dt.strftime("%b")

    #get the month column to string (not essential)
    df_sa_kit["month"] = df_sa_kit["month"].astype("str")

    kit_list = df_sa_kit.columns.values.tolist()

    df_sa_kit["Kitsap: Total NonFarm"] = df_sa_kit["Total Nonfarm"] 
    df_sa_kit["Kitsap: Const./Res"] = df_sa_kit[kit_list[3]]
    #add new column for FIRE with zero as values with index aligned with the dataframe
    df_sa_kit["Kitsap: FIRE"] = pd.Series([0 for x in range(len(df_sa_kit.index))], index=df_sa_kit.index)
    df_sa_kit["Kitsap: Manufacturing"] = df_sa_kit[kit_list[4]]
    df_sa_kit["Kitsap: Retail"] = df_sa_kit[kit_list[8]]
    df_sa_kit["Kitsap: Service"] = df_sa_kit[kit_list[5]] - df_sa_kit[kit_list[7]] - df_sa_kit[kit_list[11]]
    df_sa_kit["Kitsap: WTU"] = df_sa_kit[kit_list[7]] - df_sa_kit[kit_list[8]]
    df_sa_kit["Kitsap: Government"] = df_sa_kit[kit_list[11]]
    df_sa_kit["Kitsap: Education"] = pd.Series([0 for x in range(len(df_sa_kit.index))], index=df_sa_kit.index)

    df_sa_kit = df_sa_kit.drop(df_sa_kit.columns[:16], axis = 1)

    if df_sa_kit["Kitsap: Total NonFarm"].isin(df_sa_kit.ix[:,3:].sum(axis=1)).all() == True:
        print "Checked: Kitsap County- All industry sectors add up to total NonFarm jobs..."
    else:
        print "\nERROR!!"
    return df_sa_kit



def sa_pie():
    df_sa_pie = pd.read_excel(os.getcwd()+"/"+ sa_filename, sheetname="Tacoma MSA", header = 1, index_col=[1])
    df_sa_pie = df_sa_pie.transpose()
    df_sa_pie['date_stamp'] = df_sa_pie.index
    df_sa_pie = df_sa_pie.ix[1:]
    df_sa_pie = df_sa_pie.reset_index(drop =True)
    df_sa_pie.columns.names = ["id"]
    df_sa_pie["year"] = df_sa_pie["date_stamp"].dt.year
    df_sa_pie["month"] = df_sa_pie["date_stamp"].dt.strftime("%b")

    #get the month column to string (not essential)
    df_sa_pie["month"] = df_sa_pie["month"].astype("str")

    pie_list = df_sa_pie.columns.values.tolist()

    df_sa_pie["Pierce: Total NonFarm"] = df_sa_pie["Total Nonfarm"] 
    df_sa_pie["Pierce: Const./Res"] = df_sa_pie[pie_list[3]] + df_sa_pie[pie_list[4]]
    df_sa_pie["Pierce: FIRE"] = df_sa_pie[pie_list[15]]
    df_sa_pie["Pierce: Manufacturing"] = df_sa_pie[pie_list[6]]
    df_sa_pie["Pierce: Retail"] = df_sa_pie[pie_list[10]]
    df_sa_pie["Pierce: Service"] = df_sa_pie[pie_list[7]] - df_sa_pie[pie_list[8]] - df_sa_pie[pie_list[15]] - df_sa_pie[pie_list[25]]                                                     
    df_sa_pie["Pierce: WTU"] = df_sa_pie[pie_list[8]] - df_sa_pie[pie_list[10]]
    df_sa_pie["Pierce: Government"] = df_sa_pie[pie_list[25]] - df_sa_pie[pie_list[28]] - df_sa_pie[pie_list[30]]
    df_sa_pie["Pierce: Education"] = df_sa_pie[pie_list[28]] + df_sa_pie[pie_list[30]]

    df_sa_pie = df_sa_pie.drop(df_sa_pie.columns[:32], axis = 1)    

    if df_sa_pie["Pierce: Total NonFarm"].isin(df_sa_pie.ix[:,3:].sum(axis=1)).all() == True:
        print "Checked: Pierce County- All industry sectors add up to total NonFarm jobs..."
    else:
        print "\nERROR!!"
    return df_sa_pie



#merge the three dataframes to produce a regional dataframe
def sa_region():
    df_sa_psrc = pd.merge(pd.merge(df_sa_kin_sno, df_sa_pie, on = ["year", "month"]), df_sa_kit, on = ["year", "month"])
    
    
    #define regional sectors and add columns to a new dataframe

    df_sa_region  = df_sa_psrc
    region_list = df_sa_region.columns.values.tolist()

    df_sa_region["Region: Const./Res"] = df_sa_region[region_list[3]] + df_sa_region[region_list[12]] + df_sa_region[region_list[21]]
    df_sa_region["Region: FIRE"] = df_sa_region[region_list[4]] + df_sa_region[region_list[13]] + df_sa_region[region_list[22]]
    df_sa_region["Region: Manufacturing"] = df_sa_region[region_list[5]] + df_sa_region[region_list[14]] + df_sa_region[region_list[23]]
    df_sa_region["Region: Retail"] = df_sa_region[region_list[6]] + df_sa_region[region_list[15]] + df_sa_region[region_list[24]]
    df_sa_region["Region: Service"] = df_sa_region[region_list[7]] + df_sa_region[region_list[16]] + df_sa_region[region_list[25]]
    df_sa_region["Region: WTU"] = df_sa_region[region_list[8]] + df_sa_region[region_list[17]] + df_sa_region[region_list[26]]
    df_sa_region["Region: Education"] = df_sa_region[region_list[10]] + df_sa_region[region_list[19]] + df_sa_region[region_list[28]]
    df_sa_region["Region: Government"] = df_sa_region[region_list[9]] + df_sa_region[region_list[18]] + df_sa_region[region_list[27]]
    df_sa_region["Region: Total NonFarm"] = df_sa_region[region_list[2]] + df_sa_region[region_list[11]] + df_sa_region[region_list[20]]
    
    
    print "Checked: Regional dataframe completed..."
    
    return df_sa_region


#delete unnecessary worksheets from the excel file

def delete_ws():
    wb_sa = load_workbook(os.getcwd()+"/"+ sa_filename)

    del_sheets = []

    for i in wb_sa.get_sheet_names():
        if i == "Seattle MSA":
            pass
        elif i == "Tacoma MSA":
            pass
        elif i == "Bremerton MSA":
            pass
        elif i == "Washington State":
            pass
        else:
            del_sheets.append(i)
    #print del_sheets
    #now delete the bad sheets
    [wb_sa.remove_sheet(wb_sa.get_sheet_by_name(sheet)) for sheet in del_sheets]
    #if you get a ValueError, this means you are incorrectly requesting a sheet

    wb_sa.save(os.getcwd()+"/"+ sa_filename)
    
    print "\nWorksheets not included in the processing been deleted..."


def data_to_workbook():
    sa_book = load_workbook(os.getcwd()+"/"+ sa_filename)
    sa_writer = pd.ExcelWriter(os.getcwd()+"/"+ sa_filename, engine='openpyxl') 
    sa_writer.book = sa_book
    sa_writer.sheets = dict((ws.title, ws) for ws in sa_book.worksheets)

    df_sa_region_total.to_excel(sa_writer, "Region_Master", index =False)

    df_sa_region.to_excel(sa_writer, "Region_Sector", index =False)

    df_sa_region_total_CM.to_excel(sa_writer, "Region_CMonth", index =False)

    df_sa_region_sector_CM.to_excel(sa_writer, "Region_Sec_CMonth", index =False)

    sa_writer.save()
    
    print "\nProcessed data been exported to the excel file located at the working directory: " + "\n" + os.getcwd() +"\n" + "\nNew Worksheets: \n1. Region_Master \n2. Region_Sector \n3. Region_CMonth \n4. Region_Sec_CMonth "


    
    
def region_yearly_plot():
    PSRC = go.Scatter(
        x=df_by_month["period"], # assign x as the dataframe column 'x'
        y=df_by_month['Region: Total NonFarm'],
        line=dict(color='rgb(230,85,13)', width=5),
        hoverinfo = "all", showlegend = False, name= "PSRC")
        
    loc_2010 = df_by_month[(df_by_month["year"]== hist_year) & (df_by_month["month"] == active_month)].index.tolist()
    for i in loc_2010: loc_2010 = i
    loc_2010_5 = df_by_month[(df_by_month["year"]== hist_year+5) & (df_by_month["month"] == active_month)].index.tolist()
    for i in loc_2010_5: loc_2010_5 = i     

    First_Marker = go.Scatter(
                    x=[df_by_month["period"].iloc[loc_2010]],
                    y=[df_by_month['Region: Total NonFarm'].iloc[loc_2010]],
                    # x=df_by_month["period"][(df_by_month["year"]== 2010) & (df_by_month["month"] == active_month)],
                    # y=df_by_month['Region: Total NonFarm'][(df_by_month["year"]== 2010) & (df_by_month["month"] == active_month)],
                    mode='markers',
                    marker=dict(color="black", size=15, symbol = "square"), 
                    text = None, 
                    hoverinfo = "none",
                    showlegend = False)

    Middle_Marker = go.Scatter(
                    x=[ df_by_month["period"].iloc[loc_2010_5]],
                    y=[df_by_month['Region: Total NonFarm'].iloc[loc_2010_5]],
                    # x= df_by_month["period"][(df_by_month["year"]== (2010+4)) & (df_by_month["month"] == active_month)],
                    # y= df_by_month['Region: Total NonFarm'][(df_by_month["year"]== (2010+4)) & (df_by_month["month"] == active_month)],
                    mode='markers',
                    marker=dict(color="black", size=15, symbol = "square"), 
                    text = None, 
                    hoverinfo = "none",
                    showlegend = False)

    Last_Marker = go.Scatter(
                    x=[ df_by_month["period"].iloc[-1]],
                    y=[df_by_month['Region: Total NonFarm'].iloc[-1]],
                    mode='markers',
                    marker=dict(color="black", size=15, symbol = "square"), 
                    hoverinfo = "none",
                    showlegend = False)

    data = [PSRC, First_Marker, Middle_Marker, Last_Marker]


    layout = go.Layout(
        title="Regional Wage & Salary Jobs " + str(hist_year)+ "-" + str(df_sa_region_total.iloc[-1,0])+"<br>(Available latest month: " + str(df_sa_region_total.iloc[-1,1]) + ", " +str(df_sa_region_total.iloc[-1,0])+')',
        titlefont = dict(color= "black", size = 30 ),
        yaxis=dict(
            showgrid=False,
            zeroline=False,
            showline=False,
            showticklabels=True,
            ticks='outside',
            tickfont = dict(size = 12)
            ),
        xaxis=dict(
            showgrid=False,
            zeroline=False,
            showline=True,
            showticklabels=True,
            linecolor='rgb(204, 204, 204)',
            linewidth=5,
            ticks='outside',
            tickcolor='rgb(204, 204, 204)',
            tickangle=270
            ),
        margin=dict(t=100, b=100,autoexpand=True)
    )
    
      
    annotations = [
        dict(
            x=df_by_month["period"].iloc[loc_2010_5],
            y=df_by_month['Region: Total NonFarm'].iloc[loc_2010_5],
           xref='x',
           yref='y',
           align = "center",
           xanchor='middle', 
           yanchor='bottom',
           arrowhead=2,
           arrowsize=1,
           arrowwidth=2,
           arrowcolor='#707070',
           ax=50,
           ay=50,
           bordercolor='#c7c7c7',
           borderwidth=3,
           borderpad=8,
           bgcolor='#ff7f0e',
           opacity=0.8,
           text=str(df_by_month["year"].iloc[loc_2010_5])+"-"+ str(df_by_month["month"].iloc[loc_2010_5])+":"+"<br>{0:.2f}".format(float(df_by_month['Region: Total NonFarm'].iloc[loc_2010_5])/1000000)+" M",
           font=dict(family = "Arial",size=16, color='black'),
           showarrow=True
            ),
    
        dict(
            x=df_by_month["period"].iloc[loc_2010],
            y=df_by_month['Region: Total NonFarm'].iloc[loc_2010],
           xref='x',
           yref='y',
           align = "center",
           xanchor='middle', 
           yanchor='bottom',
           arrowhead=2,
           arrowsize=1,
           arrowwidth=2,
           arrowcolor='#707070',
           ax=40,
           ay=-65,
           bordercolor='#c7c7c7',
           borderwidth=3,
           borderpad=8,
           bgcolor='#ff7f0e',
           opacity=0.8,
           text=str(df_by_month["year"].iloc[loc_2010])+"-"+ str(df_by_month["month"].iloc[loc_2010])+":"+"<br>{0:.2f}".format(float(df_by_month['Region: Total NonFarm'].iloc[loc_2010])/1000000)+" M",
           font=dict(family = "Arial",size=16, color='black'),
           showarrow=True
            ),
    
       dict(
            x=df_by_month["period"].iloc[-1],
            #x=2,
            y=df_by_month['Region: Total NonFarm'].iloc[-1],
           # y =2,
           xref='x',
           yref='y',
           align = "center",
           xanchor='middle', 
           yanchor='bottom',
           arrowhead=2,
           arrowsize=1,
           arrowwidth=2,
           arrowcolor='#707070',
           ax=50,
           ay=-50,
           bordercolor='#c7c7c7',
           borderwidth=3,
           borderpad=8,
           bgcolor='#ff7f0e',
           opacity=0.8,
           text=str(df_by_month["year"].iloc[-1])+"-"+ str(df_by_month["month"].iloc[-1])+":"+"<br>{0:.2f}".format(float(df_by_month['Region: Total NonFarm'].iloc[-1])/1000000)+"M",
           font=dict(family = "Arial",size=16, color='black'),
           showarrow=True
            ),
       dict(xref='paper', yref='paper', x=0.5, y=-0.125,
                              xanchor='center', yanchor='bottom',
                              text='Source: Wasington Employment Estimates (Seasonally Adjusted Series): Monthly Estimates of Nonfarm Employment; ' +
                                        "Available latest month: "+str(df_by_month.iloc[-1,1])+", "+str(df_by_month.iloc[-1,0])+"; "+
                                   'Date: '+ time.strftime("%m%d%Y"),
                              font=dict(family='Arial',
                                        size=12,
                                        color='rgb(150,150,150)'),
                              showarrow=False,)
    ]



    layout['annotations'] = annotations
    fig = go.Figure(data = data, layout = layout)
    plotly.offline.plot(fig, filename = "1. Region_10-16"+"_"+time.strftime("%m%d%Y")+".html")    
    

def region_sectors_plot():
    Const_Res = go.Scatter(x=df_region_sec_line["period"], y=df_region_sec_line["Region: Const./Res"], name= "Const./Res")
    FIRE = go.Scatter(x=df_region_sec_line["period"], y=df_region_sec_line["Region: FIRE"], name= "FIRE" )
    Manufacturing = go.Scatter(x=df_region_sec_line["period"], y=df_region_sec_line["Region: Manufacturing"], name= "Manufacturing")
    Retail = go.Scatter(x=df_region_sec_line["period"], y=df_region_sec_line["Region: Retail"], name= "Retail")
    Service = go.Scatter(x=df_region_sec_line["period"], y=df_region_sec_line["Region: Service"], name = "Service")
    WTU = go.Scatter(x=df_region_sec_line["period"], y=df_region_sec_line["Region: WTU"], name = "WTU")
    Government = go.Scatter(x=df_region_sec_line["period"], y=df_region_sec_line["Region: Government"], name = "Government")
    Education = go.Scatter(x=df_region_sec_line["period"], y=df_region_sec_line["Region: Education"], name= "Education")

    fig = tools.make_subplots(rows=2, cols=4, subplot_titles=('Const./Res', 'FIRE', 'Manufacturing', 'Retail', 'Service', 'WTU', 'Government', 'Education'), 
                          )

    fig.append_trace(Const_Res, 1, 1)
    fig.append_trace(FIRE, 1, 2)
    fig.append_trace(Manufacturing, 1, 3)
    fig.append_trace(Retail, 1, 4)
    fig.append_trace(Service, 2, 1)
    fig.append_trace(WTU, 2, 2)
    fig.append_trace(Government, 2, 3)
    fig.append_trace(Education, 2, 4)

    fig['layout'].update(title="Regional Wage & Salary Jobs " + str(hist_year)+ "-" + str(df_sa_region_total.iloc[-1,0])+":" +
                                                  ' PSRC Industries; (Available latest month: ' + str(df_region_sec_line.iloc[-1,1])+", "+str(df_region_sec_line.iloc[-1,0])+')', 
                                     showlegend = False, plot_bgcolor='rgba(245, 246, 249, 1)', titlefont = dict(color= "black", size = 25 ), margin=dict(t=100, b=75,autoexpand=True), 
                )

    plotly.offline.plot(fig, filename = "2. Region_10-16_Sectors"+"_"+time.strftime("%m%d%Y")+".html")



def monthly_change_plot():
    Sector_Change = go.Bar(
        x=["Const./Res", "FIRE", "Manufacturing", "Retail", "Services", "WTU", "Education", "Government", "PSRC Region"], # assign x as the dataframe column 'x'
        y=df_region_sec_subs.ix[0:-1,0],
        marker = dict(color = ['rgb(90,180,172)','rgb(90,180,172)','rgb(90,180,172)','rgb(90,180,172)','rgb(90,180,172)','rgb(90,180,172)','rgb(90,180,172)',
                               'rgb(90,180,172)','rgb(241,105,19)'], 
                      line = dict(color= ['rgb(1,102,94)','rgb(1,102,94)','rgb(1,102,94)','rgb(1,102,94)', 'rgb(1,102,94)','rgb(1,102,94)','rgb(1,102,94)',
                                          'rgb(1,102,94)','rgb(217,72,1)' ], width = 2) ),
        hoverinfo= "x+y"
    )
    
    Region_Change = go.Bar(
        x=["King & Snohomish Counties", "Pierce County", "Kitsap County", "PSRC Region"], # assign x as the dataframe column 'x'
        y=df_region_subs[0],
        marker = dict(color = ['rgb(128,125,186)','rgb(65,171,93)', 'rgb(29,145,192)','rgb(241,105,19)'], 
                      line = dict(color= ['rgb(106,81,163)','rgb(35,132,67)','rgb(0,90,50)','rgb(217,72,1)' ], width = 2) ),
    
       hoverinfo = "x+y"
    )

    title_region = ('Monthly Wage & Salary Jobs: Change in Counties, '+ str(df_sa_region_total.iloc[-2,1])+"-"+str(df_sa_region_total.iloc[-2,0])+
                                                " to " +  str(df_sa_region_total.iloc[-1,1])+"-"+str(df_sa_region_total.iloc[-1,0]))
    title_sec=('Monthly Wage & Salary Jobs: Change in Industries, '+ str(df_sa_region_total.iloc[-2,1])+"-"+str(df_sa_region_total.iloc[-2,0])+
                                                " to " +  str(df_sa_region_total.iloc[-1,1])+"-"+str(df_sa_region_total.iloc[-1,0]))

    fig = tools.make_subplots(rows=1, cols=2, subplot_titles=(title_region, title_sec))


    fig.append_trace(Region_Change, 1, 1)
    fig.append_trace(Sector_Change, 1,2)
    
    fig['layout'].update(title = ('Changes in Wage & Salary Jobs: '+ str(df_sa_region_total.iloc[-2,1])+"-"+str(df_sa_region_total.iloc[-2,0])+
                                                " to " +  str(df_sa_region_total.iloc[-1,1])+"-"+str(df_sa_region_total.iloc[-1,0])),
                                             showlegend = False, plot_bgcolor='rgba(245, 246, 249, 1)',  titlefont = dict(color= "black", size = 25 ))
    plotly.offline.plot(fig, filename = "3. Region_Monthly_Change"+"_"+time.strftime("%m%d%Y")+".html")

    

process_start = time.time()
    
download_esd_sa()
check_worksheet()

df_sa_kin_sno = sa_kin_sno()
df_sa_kit = sa_kit()
df_sa_pie = sa_pie()
df_sa_region = sa_region()


#Query: just the three MSA and regional total
df_sa_region_total = df_sa_region[[0,1, 2, 11,20,-1]]
#Query: just the three MSA and regional total for current month
df_sa_region_total_CM = df_sa_region[[0,1, 2, 11,20,-1]][df_sa_region.month== active_month]
#Query: just the three MSA and regional sectors for current month
df_sa_region_sector_CM = df_sa_region[df_sa_region.month== active_month]    

delete_ws() 
    
data_to_workbook()

print "\nProcessing data for plotting..."

#setup data for regional yearly (2010-Current) plot:

df_by_month = df_sa_region_total[df_sa_region_total["year"]>= hist_year]
df_by_month = df_by_month.reset_index(drop=True)
df_by_month["period"] = df_by_month["year"].map(str) + "-"+df_by_month["month"]
df_by_month["period"] = pd.to_datetime(df_by_month["period"])

#setup data for regional PSRC sectors line plots:

ls_reg_sec = list(range(29,38))
ls_reg_sec.append(0)
ls_reg_sec.append(1)
ls_reg_sec.sort()


df_region_sec_line = df_sa_region[ls_reg_sec]
df_region_sec_line = df_region_sec_line[(df_region_sec_line["year"]>= hist_year) ]
df_region_sec_line["period"] = df_region_sec_line["year"].map(str) + "-"+df_region_sec_line["month"]
df_region_sec_line["period"] = pd.to_datetime(df_region_sec_line['period'])
df_region_sec_line = df_region_sec_line.reset_index(drop=True)

#setup data for monthly change (region and PSRC Sectors):

df_region_subs= pd.DataFrame(df_sa_region_total.iloc[-1,2:]-df_sa_region_total.iloc[-2,2:])
df_region_sec_bar = df_sa_region[ls_reg_sec]
df_region_sec_subs= pd.DataFrame(df_region_sec_bar.iloc[-1,2:]-df_region_sec_bar.iloc[-2,2:])


region_yearly_plot()
region_sectors_plot()
monthly_change_plot()

print "\nThree plot reports are created and deployed to system's defaut web-browser. Also plots been exported as .html files to the working directory: " + "\n" + os.getcwd() +"\n" + "\nNew HTML files:" + "\n1. Region_10-16"+"_"+time.strftime("%m%d%Y")+".html" + "\n2. Region_10-16_Sectors"+"_"+time.strftime("%m%d%Y")+".html" + "\n3. Region_Monthly_Change"+"_"+time.strftime("%m%d%Y")+".html"

process_end = time.time()

processtime = process_end - process_start    

print "\nProcessing Completed! \nTotal Processing Time: " + "{0: .3}".format(processtime) + " Sec."





###############

#Date Time:

#http://stackoverflow.com/questions/311627/how-to-print-date-in-a-regular-format-in-python

# Time: http://stackoverflow.com/questions/7370801/measure-time-elapsed-in-python
# Time: http://stackoverflow.com/questions/311627/how-to-print-date-in-a-regular-format-in-python

# Date: http://stackoverflow.com/questions/25146121/extracting-just-month-and-year-from-pandas-datetime-column-python

# Date time to String: http://stackoverflow.com/questions/22276503/how-to-i-change-data-type-of-pandas-data-frame-to-string-with-a-defined-format
# Get-Current-Directory: 
# http://stackoverflow.com/questions/5137497/find-current-directory-and-files-directory 

# http://stackoverflow.com/questions/3430372/how-to-get-full-path-of-current-files-directory-in-python

# Download file using urllib:

# http://stackoverflow.com/questions/19602931/basic-http-file-downloading-and-saving-to-disk-in-python

# http://stackoverflow.com/questions/22676/how-do-i-download-a-file-over-http-using-python


#Delete rows in dataframe:

# http://stackoverflow.com/questions/16396903/delete-first-three-rows-of-a-dataframe-in-pandas

#delete columns in dataframe:

# http://stackoverflow.com/questions/13411544/delete-column-from-pandas-dataframe

#Rename Columns: http://stackoverflow.com/questions/11346283/renaming-columns-in-pandas

#Sum rows in the dataframe:

# http://stackoverflow.com/questions/25748683/python-pandas-sum-dataframe-rows-for-given-columns

#Merge multiple dataframes on columns:

# http://stackoverflow.com/questions/23668427/pandas-joining-multiple-dataframes-on-columns

# http://stackoverflow.com/questions/18792918/pandas-combining-2-data-frames-join-on-a-common-column

#Write dataframe to existing excel file:

# http://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data

#delete worksheets from existing excel file:
# http://stackoverflow.com/questions/24989506/python-2-7-i-cant-remove-worksheets-using-xlwt-workbook-workbook-worksheets 

#get index of cell which matches certain column values:
# http://stackoverflow.com/questions/21800169/python-pandas-get-index-of-rows-which-column-matches-certain-value